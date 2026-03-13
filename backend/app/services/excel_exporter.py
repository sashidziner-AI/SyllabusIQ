import logging
import os
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session

from app.config import settings
from app.models.export_job import ExportJob, ExportStatus
from app.models.question import MCQuestion
from app.models.nos_unit import NOSUnit
from app.models.criterion import PerformanceCriterion

logger = logging.getLogger(__name__)

HEADERS = [
    "#", "Question", "Option A", "Option B", "Option C", "Option D",
    "Correct Answer", "Explanation", "Source Page", "NOS Unit",
    "Criterion", "Difficulty",
]


def create_export(
    db: Session,
    user_id: int,
    filter_criteria: dict,
) -> ExportJob:
    """Create an Excel export of questions based on filter criteria."""
    filename = f"syllabus_iq_export_{uuid.uuid4().hex[:8]}.xlsx"
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    file_path = os.path.join(settings.EXPORT_DIR, filename)

    job = ExportJob(
        user_id=user_id,
        filename=filename,
        format="xlsx",
        filter_criteria=filter_criteria,
        status=ExportStatus.processing,
        file_path=file_path,
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    try:
        query = db.query(MCQuestion).filter(MCQuestion.user_id == user_id)

        doc_ids = filter_criteria.get("document_ids")
        if doc_ids:
            query = query.filter(MCQuestion.document_id.in_(doc_ids))

        difficulty_levels = filter_criteria.get("difficulty_levels")
        if difficulty_levels:
            query = query.filter(MCQuestion.difficulty_level.in_(difficulty_levels))

        nos_unit_ids = filter_criteria.get("nos_unit_ids")
        if nos_unit_ids:
            query = query.filter(MCQuestion.nos_unit_id.in_(nos_unit_ids))

        questions = query.order_by(MCQuestion.id).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "MCQ Export"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")

        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_num, q in enumerate(questions, 2):
            nos_unit_name = ""
            if q.nos_unit_id:
                nos_unit = db.query(NOSUnit).filter(NOSUnit.id == q.nos_unit_id).first()
                if nos_unit:
                    nos_unit_name = f"{nos_unit.unit_code} - {nos_unit.unit_title}"

            criterion_name = ""
            if q.criterion_id:
                criterion = db.query(PerformanceCriterion).filter(
                    PerformanceCriterion.id == q.criterion_id
                ).first()
                if criterion:
                    criterion_name = f"{criterion.criterion_code}: {criterion.criterion_text}"

            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=q.question_text)
            ws.cell(row=row_num, column=3, value=q.option_a)
            ws.cell(row=row_num, column=4, value=q.option_b)
            ws.cell(row=row_num, column=5, value=q.option_c)
            ws.cell(row=row_num, column=6, value=q.option_d)
            ws.cell(row=row_num, column=7, value=q.correct_option)
            ws.cell(row=row_num, column=8, value=q.explanation)
            ws.cell(row=row_num, column=9, value=q.source_page_reference or "")
            ws.cell(row=row_num, column=10, value=nos_unit_name)
            ws.cell(row=row_num, column=11, value=criterion_name)
            ws.cell(row=row_num, column=12, value=q.difficulty_level.value if q.difficulty_level else "")

        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        wb.save(file_path)

        job.status = ExportStatus.completed
        job.row_count = len(questions)
        job.completed_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(job)
        logger.info("Export completed: %s with %d questions", filename, len(questions))

    except Exception as e:
        job.status = ExportStatus.failed
        db.commit()
        logger.error("Export failed: %s", str(e))
        raise

    return job
